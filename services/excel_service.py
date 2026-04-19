import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

def gerar_excel(dados, nome_arquivo):
    df = pd.DataFrame(dados)

    if "id" in df.columns:
        df.sort_values(by="id", inplace=True)

    try:
        df.to_excel(nome_arquivo, index=False)

        wb = load_workbook(nome_arquivo)
        ws = wb.active

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        for cell in ws[1]:
            cell.font = Font(bold=True)

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        wb.save(nome_arquivo)
        print(f"✅ Arquivo '{nome_arquivo}' gerado com sucesso!")

    except Exception as e:
        print(f"❌ Erro ao salvar o arquivo: {e}")